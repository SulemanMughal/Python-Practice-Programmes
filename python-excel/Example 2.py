# Navigate through different cells
import openpyxl

# Open an excel file in current directory
wb = openpyxl.load_workbook('file_example_XLSX_10.xlsx')

# Get active sheet
sheet = wb.active

print("Active Sheet : \n", sheet)

# Select a cell
print("A1" , sheet['A1'])
print( "A1 value" , sheet['A1'].value)
c = sheet['A1']
print("Row " + str(c.row) + " Column " +  str(c.column)  + ' is ' + str(c.value))
print('Cell ' + c.coordinate + ' is ' + str(c.value))
print(sheet['C1'].value)
print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)
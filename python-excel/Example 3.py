# Determine size of sheet
import openpyxl

# Open an excel file in current directory
wb = openpyxl.load_workbook('file_example_XLSX_10.xlsx')

# Get active sheet
sheet = wb.active

# Get maximum row number
row_count = sheet.max_row

# Get maximum column number
column_count = sheet.max_column

print(row_count, column_count, sep="\n")
# Converting Between Column Letters and Numbers
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
# Open an excel file in current directory
wb = openpyxl.load_workbook('file_example_XLSX_10.xlsx')

# Get active sheet
sheet = wb.active

print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(900))
print(column_index_from_string('A'))
print(column_index_from_string('AA'))